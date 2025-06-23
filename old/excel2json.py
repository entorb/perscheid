"""
Reads excel file, exports as json list.
"""  # noqa: INP001

import json
import urllib.parse
from pathlib import Path

import openpyxl

FILE_EXCEL = Path("datenbank.xlsx")
FILE_JSON = Path("datenbank.json")

workbook = openpyxl.load_workbook(
    FILE_EXCEL,
    data_only=True,
)  # data_only : read values instead of formulas

sheet = workbook["datenbank"]

num_cols = 5

# header
# cspell:disable-next-line
header = ("Originalnummer", "Stichworte", "A-Nummern", "Buch", "PostkartenNr")

row = 1
# ensure the columns are not moved
for col in range(1, 1 + num_cols):
    cell = sheet.cell(row=row, column=col)  # index start here with 1
    # print(cell.value)
    assert cell.value == header[col - 1]


def gen_url(text: str) -> str:  # noqa: D103
    s = text
    s = s.replace("?", " ")
    s = s.replace(".", " ")
    s = s.replace(",", " ")
    s = s.replace("!", " ")
    s = s.replace("-", " ")
    s = s.replace("/", " ")
    s = s.strip()  # trim spaces from both sides, rstrip for right only

    # replace all (multiple) whitespaces by single space ' '
    s = " ".join(s.split())

    # url encoding of special chars
    s = urllib.parse.quote_plus(s)
    # print(s)

    url = f"https://www.google.de/search?q=Perscheid+{s}&hl=de&tbm=isch"
    return url


list_of_cartoons = []

row = 2
my_id = 0
while my_id is not None:
    my_id = sheet.cell(row=row, column=1).value
    if my_id is None:
        break
    text = str(sheet.cell(row=row, column=2).value)
    url = gen_url(text)
    d = {
        # cspell:disable
        "Originalnummer": my_id,
        "Stichworte": text,
        "A-Nummern": sheet.cell(row=row, column=3).value,
        "Buch": sheet.cell(row=row, column=4).value,
        "PostkartenNr": sheet.cell(row=row, column=5).value,
        "URL": url,
        # cspell:enable
    }
    if my_id is not None:
        list_of_cartoons.append(d)

    row += 1


with FILE_JSON.open(mode="w", encoding="utf-8", newline="\n") as fh:
    json.dump(list_of_cartoons, fh, ensure_ascii=False, sort_keys=False, indent=1)
